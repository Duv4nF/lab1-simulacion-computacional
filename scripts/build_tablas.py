# -*- coding: utf-8 -*-
"""Genera tablas LaTeX, figuras y archivos .xlsx de la Practica 1."""
import os, io, json
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from simcore import (make_runs, confidence, MEASURE_KEYS, MEASURE_ES,
                     MEASURE_SHORT, N_RUNS, SEED)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "Entrega_Lab1")
TEX = os.path.join(OUT, "Informe_LaTeX")
os.makedirs(TEX, exist_ok=True)

BLUE, ORANGE = "#2a78d6", "#eb6834"
INK, INK2, MUTED, GRID, AXIS = "#0b0b0b", "#52514e", "#898781", "#e1e0d9", "#c3c2b7"

DATA = {20: make_runs(20), 200: make_runs(200)}
STATS = {}
for N, runs in DATA.items():
    STATS[N] = {}
    for k in MEASURE_KEYS:
        sample = [r["meas"][k] for r in runs]
        STATS[N][k] = {"sample": sample,
                       95: confidence(sample, 0.05),
                       99: confidence(sample, 0.01)}


def w(name, text):
    with io.open(os.path.join(TEX, name), "w", encoding="utf-8") as f:
        f.write(text)
    print("  escrito", name)


def num(x, d=4):
    return ("%." + str(d) + "f") % x


# ---------------------------------------------------------------- tabla detallada
def tabla_detalle(N, run_idx=0):
    t = DATA[N][run_idx]["table"]
    m = DATA[N][run_idx]["meas"]
    head = (r"\makecell{(1)\\Cliente} & \makecell{(2)\\Tiempo entre\\llegadas} & "
            r"\makecell{(3)\\Hora de\\llegada} & \makecell{(4)\\Tiempo de\\servicio} & "
            r"\makecell{(5)\\Inicio del\\servicio} & \makecell{(6)\\Fin del\\servicio} & "
            r"\makecell{(7)\\Tiempo en\\el sistema} & \makecell{(8)\\Tiempo\\ocioso} & "
            r"\makecell{(9)\\Tiempo en\\cola} \\")
    L = []
    L.append(r"\begingroup\small\setlength{\tabcolsep}{4pt}")
    L.append(r"\begin{longtable}{@{}rrrrrrrrr@{}}")
    L.append(r"\caption{Simulaci\'on \emph{ad hoc} de la fila del banco con %d clientes "
             r"(corrida %d).}\label{tab:sim%d} \\" % (N, run_idx + 1, N))
    L.append(r"\toprule")
    L.append(head)
    L.append(r"\midrule \endfirsthead")
    L.append(r"\multicolumn{9}{@{}l}{\footnotesize\itshape Tabla \thetable{} "
             r"(continuaci\'on)}\\[2pt]")
    L.append(r"\toprule")
    L.append(head)
    L.append(r"\midrule \endhead")
    L.append(r"\midrule \multicolumn{9}{r@{}}{\footnotesize\itshape Contin\'ua en la "
             r"p\'agina siguiente}\\ \endfoot")
    L.append(r"\bottomrule \endlastfoot")
    for i in range(N):
        tba = "--" if i == 0 else str(int(t["tba"][i]))
        L.append("%d & %s & %d & %d & %d & %d & %d & %d & %d \\\\" % (
            i + 1, tba, t["arrival"][i], t["service"][i], t["begin"][i],
            t["end"][i], t["system"][i], t["idle"][i], t["queue"][i]))
    L.append(r"\midrule")
    L.append(r"\textbf{Suma} & & & & & & \textbf{%d} & \textbf{%d} & \textbf{%d} \\"
             % (m["_sum_system"], m["_sum_idle"], m["_sum_queue"]))
    L.append(r"\end{longtable}")
    L.append(r"\endgroup")
    return "\n".join(L)


# ---------------------------------------------------------------- medidas
def tabla_medidas(N, run_idx=0):
    m = DATA[N][run_idx]["meas"]
    n_w = m["_n_waiters"]
    filas = [
        ("Tiempo promedio en el sistema",
         r"$\sum(7)/N = %d/%d$" % (m["_sum_system"], N), num(m["avg_system"]), "min"),
        ("Porcentaje de tiempo ocioso",
         r"$\sum(8)/T = %d/%d$" % (m["_sum_idle"], m["_total_time"]),
         num(m["pct_idle"]) + r"\ (%s\%%)" % num(100 * m["pct_idle"], 2), "---"),
        ("Tiempo de espera promedio por cliente",
         r"$\sum(9)/N = %d/%d$" % (m["_sum_queue"], N), num(m["avg_wait_all"]), "min"),
        ("Fracci\\'on de clientes que esper\\'o",
         r"$N_{e}/N = %d/%d$" % (n_w, N), num(m["frac_wait"]), "---"),
        ("Tiempo de espera promedio de quienes esperaron",
         r"$\sum(9)/N_{e} = %d/%d$" % (m["_sum_queue"], n_w),
         num(m["avg_wait_waiters"]), "min"),
    ]
    L = [r"\begin{table}[H]", r"\centering",
         r"\caption{Medidas de desempe\~no de la corrida %d con %d clientes.}" % (run_idx + 1, N),
         r"\label{tab:meas%d}" % N, r"\small\setlength{\tabcolsep}{5pt}",
         r"\begin{tabular}{@{}llrl@{}}", r"\toprule",
         r"Medida de desempe\~no & C\'alculo & Valor & Unidad \\", r"\midrule"]
    for a, b, c, d in filas:
        L.append("%s & %s & %s & %s \\\\" % (a, b, c, d))
    L += [r"\bottomrule", r"\end{tabular}", r"\end{table}"]
    return "\n".join(L)


# ---------------------------------------------------------------- 10 corridas
def tabla_corridas(N):
    L = [r"\begin{table}[H]", r"\centering",
         r"\caption{Medidas de desempe\~no de las %d corridas con %d clientes.}" % (N_RUNS, N),
         r"\label{tab:runs%d}" % N, r"\small",
         r"\begin{tabular}{@{}rrrrrr@{}}", r"\toprule",
         r"\makecell{Corrida} & \makecell{Tiempo prom.\\en el sistema\\(min)} & "
         r"\makecell{Tiempo\\ocioso} & \makecell{Espera prom.\\por cliente\\(min)} & "
         r"\makecell{Fracci\'on\\que esper\'o} & \makecell{Espera prom. de\\quienes esperaron\\(min)} \\",
         r"\midrule"]
    for i, r in enumerate(DATA[N], 1):
        m = r["meas"]
        L.append("%d & %s & %s & %s & %s & %s \\\\" % (
            i, num(m["avg_system"], 2), num(m["pct_idle"]), num(m["avg_wait_all"], 3),
            num(m["frac_wait"], 3), num(m["avg_wait_waiters"], 3)))
    L.append(r"\midrule")
    L.append(r"\textbf{Media} & %s & %s & %s & %s & %s \\" % tuple(
        num(STATS[N][k][95]["mean"], 4) for k in MEASURE_KEYS))
    L.append(r"\textbf{Desv. est. $S$} & %s & %s & %s & %s & %s \\" % tuple(
        num(STATS[N][k][95]["s"], 4) for k in MEASURE_KEYS))
    L += [r"\bottomrule", r"\end{tabular}", r"\end{table}"]
    return "\n".join(L)


# ---------------------------------------------------------------- intervalos
def tabla_ic(N):
    L = [r"\begin{table}[H]", r"\centering",
         r"\caption{Intervalos de confianza al 95\,\%% y 99\,\%% de las medidas de "
         r"desempe\~no ($n=%d$ corridas de %d clientes).}" % (N_RUNS, N),
         r"\label{tab:ci%d}" % N, r"\footnotesize",
         r"\begin{tabular}{@{}lcrrrrrr@{}}", r"\toprule",
         r"Medida & Conf. & $\bar{X}$ & $S$ & $t_{n-1,1-\alpha/2}$ & $h$ & LI & LS \\",
         r"\midrule"]
    for j, k in enumerate(MEASURE_KEYS):
        for c in (95, 99):
            d = STATS[N][k][c]
            name = MEASURE_SHORT[k].replace("Fraccion", "Fracci\\'on").replace(
                "espero", "esper\\'o")
            L.append(r"%s & %d\,\%% & %s & %s & %s & %s & %s & %s \\" % (
                name if c == 95 else "", c, num(d["mean"]), num(d["s"]),
                num(d["t"]), num(d["h"]), num(d["lo"]), num(d["hi"])))
        if j < len(MEASURE_KEYS) - 1:
            L.append(r"\addlinespace")
    L += [r"\bottomrule", r"\end{tabular}",
          r"\begin{minipage}{\textwidth}\vspace{2pt}\footnotesize",
          r"LI y LS: l\'imites inferior y superior del intervalo. "
          r"$t_{9,0.975}=2{,}2622$ y $t_{9,0.995}=3{,}2498$.",
          r"\end{minipage}", r"\end{table}"]
    return "\n".join(L)


# ---------------------------------------------------------------- resumen comparativo
def tabla_resumen():
    L = [r"\begin{table}[H]", r"\centering",
         r"\caption{Comparaci\'on de la desviaci\'on est\'andar muestral y del rango del "
         r"intervalo de confianza ($2h$) entre los dos escenarios.}",
         r"\label{tab:resumen}", r"\footnotesize",
         r"\begin{tabular}{@{}lrrrrrr@{}}", r"\toprule",
         r"\multirow{2}{*}{Medida} & \multicolumn{2}{c}{$S$} & "
         r"\multicolumn{2}{c}{Rango IC 95\,\% ($2h$)} & \multicolumn{2}{c}{Rango IC 99\,\% ($2h$)} \\",
         r"\cmidrule(lr){2-3}\cmidrule(lr){4-5}\cmidrule(l){6-7}",
         r" & 20 cl. & 200 cl. & 20 cl. & 200 cl. & 20 cl. & 200 cl. \\", r"\midrule"]
    for k in MEASURE_KEYS:
        a20, a200 = STATS[20][k], STATS[200][k]
        name = MEASURE_SHORT[k].replace("Fraccion", "Fracci\\'on").replace("espero", "esper\\'o")
        vals = [a20[95]["s"], a200[95]["s"], a20[95]["range"], a200[95]["range"],
                a20[99]["range"], a200[99]["range"]]
        L.append("%s & %s \\\\" % (name, " & ".join(num(v) for v in vals)))
    L += [r"\bottomrule", r"\end{tabular}", r"\end{table}"]
    return "\n".join(L)


# ---------------------------------------------------------------- precision relativa
def tabla_cv():
    L = [r"\begin{table}[H]", r"\centering",
         r"\caption{Precisi\'on relativa: coeficiente de variaci\'on ($S/\bar{X}$) y "
         r"semiamplitud relativa al 95\,\% ($h/\bar{X}$), en porcentaje.}",
         r"\label{tab:cv}", r"\footnotesize",
         r"\begin{tabular}{@{}lrrrrr@{}}", r"\toprule",
         r"\multirow{2}{*}{Medida} & \multicolumn{2}{c}{$S/\bar{X}$ (\%)} & "
         r"\multicolumn{2}{c}{$h/\bar{X}$ al 95\,\% (\%)} & \multirow{2}{*}{\makecell{Reducci\'on\\de $S$}} \\",
         r"\cmidrule(lr){2-3}\cmidrule(lr){4-5}",
         r" & 20 cl. & 200 cl. & 20 cl. & 200 cl. & \\", r"\midrule"]
    for k in MEASURE_KEYS:
        a, b = STATS[20][k][95], STATS[200][k][95]
        name = MEASURE_SHORT[k].replace("Fraccion", "Fracci\\'on").replace("espero", "esper\\'o")
        L.append("%s & %s & %s & %s & %s & %s\\,\\%% \\\\" % (
            name, num(100 * a["s"] / a["mean"], 2), num(100 * b["s"] / b["mean"], 2),
            num(100 * a["h"] / a["mean"], 2), num(100 * b["h"] / b["mean"], 2),
            num(100 * (1 - b["s"] / a["s"]), 1)))
    L += [r"\bottomrule", r"\end{tabular}", r"\end{table}"]
    return "\n".join(L)


# ---------------------------------------------------------------- validacion generador
def tabla_validacion():
    from scipy import stats as st
    from simcore import N_RUNS as NR
    L = [r"\begin{table}[H]", r"\centering",
         r"\caption{Verificaci\'on de los generadores: prueba $\chi^2$ de bondad de ajuste "
         r"a la distribuci\'on uniforme discreta.}", r"\label{tab:valgen}", r"\footnotesize",
         r"\begin{tabular}{@{}llrrrrr@{}}", r"\toprule",
         r"Escenario & Variable & \makecell{Valores\\generados} & "
         r"\makecell{Media\\observada} & \makecell{Media\\te\'orica} & $\chi^2$ & valor $p$ \\",
         r"\midrule"]
    for N in (20, 200):
        rt = np.random.default_rng(SEED)
        rs = np.random.default_rng(SEED + 1)
        T, S = [], []
        for _ in range(NR):
            T.append(rt.integers(1, 11, size=N)[1:])
            S.append(rs.integers(1, 7, size=N))
        T, S = np.concatenate(T), np.concatenate(S)
        for etiq, dat, k, teor in (("Tiempo entre llegadas $\\sim U\\{1,10\\}$", T, 10, 5.5),
                                   ("Tiempo de servicio $\\sim U\\{1,6\\}$", S, 6, 3.5)):
            f = np.bincount(dat, minlength=k + 1)[1:k + 1]
            c = st.chisquare(f)
            L.append(r"%s & %s & %d & %s & %s & %s & %s \\" % (
                (r"%d clientes" % N) if teor == 5.5 else "", etiq, len(dat),
                num(dat.mean(), 3), num(teor, 3), num(c.statistic, 3), num(c.pvalue, 3)))
        if N == 20:
            L.append(r"\addlinespace")
    L += [r"\bottomrule", r"\end{tabular}",
          r"\begin{minipage}{\textwidth}\vspace{2pt}\footnotesize",
          r"Hip\'otesis nula: los valores provienen de la distribuci\'on uniforme discreta "
          r"indicada. Con $p>0{,}05$ en los cuatro casos no hay evidencia para rechazarla.",
          r"\end{minipage}", r"\end{table}"]
    return "\n".join(L)


# ---------------------------------------------------------------- figuras
def estilo(ax):
    ax.set_facecolor("#fcfcfb")
    for s in ("top", "right"):
        ax.spines[s].set_visible(False)
    for s in ("left", "bottom"):
        ax.spines[s].set_color(AXIS)
        ax.spines[s].set_linewidth(0.8)
    ax.tick_params(colors=MUTED, labelsize=8, length=3, width=0.8)
    for lb in ax.get_xticklabels() + ax.get_yticklabels():
        lb.set_color(INK2)


def figura_dispersion():
    fig, ax = plt.subplots(figsize=(6.4, 3.1))
    fig.patch.set_facecolor("#fcfcfb")
    estilo(ax)
    ax.grid(axis="x", color=GRID, linewidth=0.7, zorder=0)
    ax.set_axisbelow(True)
    rng = np.random.default_rng(7)
    for row, (N, color) in enumerate([(200, ORANGE), (20, BLUE)]):
        s = np.array(STATS[N]["avg_system"]["sample"])
        d = STATS[N]["avg_system"][95]
        jit = rng.uniform(-0.10, 0.10, len(s))
        ax.hlines(row, d["lo"], d["hi"], color=color, linewidth=8, alpha=0.20,
                  zorder=2, capstyle="round")
        ax.scatter(s, row + jit, s=40, color=color, edgecolor="#fcfcfb",
                   linewidth=1.2, zorder=4)
        ax.scatter([d["mean"]], [row], s=115, marker="D", color=color,
                   edgecolor="#fcfcfb", linewidth=1.6, zorder=5)
        ax.annotate(r"$\bar{X}=%.2f$" % d["mean"], (d["mean"], row + 0.27),
                    ha="center", fontsize=8.5, color=INK)
        ax.annotate("IC 95 %%: %.2f – %.2f" % (d["lo"], d["hi"]),
                    (d["mean"], row - 0.30), ha="center", va="top",
                    fontsize=7.8, color=MUTED)
    ax.set_yticks([0, 1])
    ax.set_yticklabels(["200 clientes\npor corrida", "20 clientes\npor corrida"], fontsize=9)
    ax.set_ylim(-0.62, 1.55)
    ax.set_xlim(3.0, 7.45)
    ax.set_xlabel("Tiempo promedio en el sistema por corrida (minutos)",
                  fontsize=8.5, color=INK2)
    ax.annotate("cada punto = una de las 10 corridas;  ◆ = media muestral;  "
                "banda = intervalo de confianza al 95 %",
                xy=(0.0, 1.06), xycoords="axes fraction", fontsize=7.6, color=MUTED)
    fig.tight_layout()
    fig.savefig(os.path.join(TEX, "fig_dispersion.pdf"), facecolor="#fcfcfb")
    plt.close(fig)
    print("  escrito fig_dispersion.pdf")


def figura_precision():
    labels = ["Tiempo prom.\nen el sistema", "Tiempo\nocioso", "Espera prom.\npor cliente",
              "Fracción\nque esperó", "Espera prom. de\nquienes esperaron"]
    v20 = [100 * STATS[20][k][95]["h"] / STATS[20][k][95]["mean"] for k in MEASURE_KEYS]
    v200 = [100 * STATS[200][k][95]["h"] / STATS[200][k][95]["mean"] for k in MEASURE_KEYS]
    y = np.arange(len(labels))
    fig, ax = plt.subplots(figsize=(6.4, 3.4))
    fig.patch.set_facecolor("#fcfcfb")
    estilo(ax)
    ax.grid(axis="x", color=GRID, linewidth=0.7)
    ax.set_axisbelow(True)
    hgt = 0.34
    ax.barh(y + hgt / 2 + 0.01, v20, height=hgt, color=BLUE, label="20 clientes por corrida")
    ax.barh(y - hgt / 2 - 0.01, v200, height=hgt, color=ORANGE, label="200 clientes por corrida")
    for yy, v in zip(y + hgt / 2 + 0.01, v20):
        ax.text(v + 0.7, yy, "%.1f%%" % v, va="center", fontsize=7.8, color=INK2)
    for yy, v in zip(y - hgt / 2 - 0.01, v200):
        ax.text(v + 0.7, yy, "%.1f%%" % v, va="center", fontsize=7.8, color=INK2)
    ax.set_yticks(y)
    ax.set_yticklabels(labels, fontsize=8)
    ax.invert_yaxis()
    ax.set_xlim(0, max(v20) * 1.18)
    ax.set_xlabel(r"Semiamplitud relativa del intervalo al 95 %  ($h/\bar{X}$, en %)",
                  fontsize=8.5, color=INK2)
    ax.legend(frameon=False, fontsize=8, loc="lower right", labelcolor=INK2,
              handletextpad=0.5)
    fig.tight_layout()
    fig.savefig(os.path.join(TEX, "fig_precision.pdf"), facecolor="#fcfcfb")
    plt.close(fig)
    print("  escrito fig_precision.pdf")


# ---------------------------------------------------------------- Excel
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="D9E2F3")
IN_FILL = PatternFill("solid", fgColor="FFF2CC")
RES_FILL = PatternFill("solid", fgColor="E2EFDA")


def hoja_ejemplo(wb, N, run_idx=0):
    ws = wb.create_sheet("Ejemplo")
    t = DATA[N][run_idx]["table"]
    heads = ["Customer", "Time Between Arrivals", "Arrival Time", "Service Time",
             "Service Begins", "Time Service Ends", "Time in System", "Idle Time",
             "Time in Queue"]
    for c, h in enumerate(heads, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = HDR_FILL
        cell.border = BORDER
    for i in range(N):
        r = i + 2
        ws.cell(row=r, column=1, value=i + 1)
        if i == 0:
            ws.cell(row=r, column=2, value="-")
            ws.cell(row=r, column=3, value=0)
            ws.cell(row=r, column=5, value="=C2")
        else:
            ws.cell(row=r, column=2, value=int(t["tba"][i]))
            ws.cell(row=r, column=3, value="=C%d+B%d" % (r - 1, r))
            ws.cell(row=r, column=5, value="=MAX(C%d,F%d)" % (r, r - 1))
        ws.cell(row=r, column=4, value=int(t["service"][i]))
        ws.cell(row=r, column=6, value="=E%d+D%d" % (r, r))
        ws.cell(row=r, column=7, value="=F%d-C%d" % (r, r))
        ws.cell(row=r, column=8, value=0 if i == 0 else "=MAX(0,C%d-F%d)" % (r, r - 1))
        ws.cell(row=r, column=9, value="=E%d-C%d" % (r, r))
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2).fill = IN_FILL
        ws.cell(row=r, column=4).fill = IN_FILL
    last = N + 1
    sr = last + 1
    ws.cell(row=sr, column=6, value="Sum").font = Font(bold=True)
    for c in (7, 8, 9):
        cl = get_column_letter(c)
        cell = ws.cell(row=sr, column=c, value="=SUM(%s2:%s%d)" % (cl, cl, last))
        cell.font = Font(bold=True)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")

    med = [("Average time in system:", "=G%d/%d" % (sr, N), "0.0000"),
           ("Percent idle time:", "=H%d/F%d" % (sr, last), "0.00%"),
           ("Average waiting time per customer:", "=I%d/%d" % (sr, N), "0.0000"),
           ("Fraction having to wait:", '=COUNTIF(I2:I%d,">0")/%d' % (last, N), "0.0000"),
           ("Average waiting time of those who waited:",
            '=I%d/COUNTIF(I2:I%d,">0")' % (sr, last), "0.0000")]
    ws.cell(row=2, column=11, value="MEDIDAS DE DESEMPEÑO").font = Font(bold=True, size=12)
    for i, (lab, f, fmt) in enumerate(med):
        r = 4 + i
        c1 = ws.cell(row=r, column=11, value=lab)
        c1.font = Font(bold=True)
        c1.border = BORDER
        c2 = ws.cell(row=r, column=12, value=f)
        c2.number_format = fmt
        c2.fill = RES_FILL
        c2.border = BORDER
        c2.alignment = Alignment(horizontal="center")
    for c, wd in zip(range(1, 13), [10, 13, 11, 11, 12, 13, 12, 10, 11, 3, 42, 12]):
        ws.column_dimensions[get_column_letter(c)].width = wd
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "A2"
    return ws


def hoja_runs(wb, N):
    ws = wb.create_sheet("10 runs")
    heads = ["Run", "Average time in system:", "Percent idle time:",
             "Average waiting time per customer:", "Fraction having to wait:",
             "Average waiting time of those who waited:"]
    for c, h in enumerate(heads, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = HDR_FILL
        cell.border = BORDER
    for i, r in enumerate(DATA[N], 1):
        m = r["meas"]
        row = i + 1
        ws.cell(row=row, column=1, value=i)
        for c, k in enumerate(MEASURE_KEYS, 2):
            cell = ws.cell(row=row, column=c, value=round(m[k], 6))
            cell.number_format = "0.0000"
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = BORDER
            ws.cell(row=row, column=c).alignment = Alignment(horizontal="center")
    base = N_RUNS + 1
    filas = [("Sample mean (X barra)", "=AVERAGE({c}2:{c}%d)" % base),
             ("Sample std. dev. (S)", "=STDEV({c}2:{c}%d)" % base),
             ("t (95%, 9 gl)", "=T.INV.2T(0.05,9)"),
             ("h 95%", "={c}%d*{c}%d/SQRT(10)"),
             ("IC 95% inferior", "={c}%d-{c}%d"),
             ("IC 95% superior", "={c}%d+{c}%d"),
             ("t (99%, 9 gl)", "=T.INV.2T(0.01,9)"),
             ("h 99%", "={c}%d*{c}%d/SQRT(10)"),
             ("IC 99% inferior", "={c}%d-{c}%d"),
             ("IC 99% superior", "={c}%d+{c}%d")]
    r_mean, r_s = base + 2, base + 3
    r_t95, r_h95 = base + 4, base + 5
    r_t99, r_h99 = base + 8, base + 9
    for i, (lab, _) in enumerate(filas):
        r = base + 2 + i
        cell = ws.cell(row=r, column=1, value=lab)
        cell.font = Font(bold=True)
        cell.border = BORDER
    for c in range(2, 7):
        cl = get_column_letter(c)
        f = {r_mean: "=AVERAGE(%s2:%s%d)" % (cl, cl, base),
             r_s: "=STDEV(%s2:%s%d)" % (cl, cl, base),
             r_t95: "=T.INV.2T(0.05,9)",
             r_h95: "=%s%d*%s%d/SQRT(10)" % (cl, r_t95, cl, r_s),
             r_t95 + 2: "=%s%d-%s%d" % (cl, r_mean, cl, r_h95),
             r_t95 + 3: "=%s%d+%s%d" % (cl, r_mean, cl, r_h95),
             r_t99: "=T.INV.2T(0.01,9)",
             r_h99: "=%s%d*%s%d/SQRT(10)" % (cl, r_t99, cl, r_s),
             r_h99 + 1: "=%s%d-%s%d" % (cl, r_mean, cl, r_h99),
             r_h99 + 2: "=%s%d+%s%d" % (cl, r_mean, cl, r_h99)}
        for r, formula in f.items():
            cell = ws.cell(row=r, column=c, value=formula)
            cell.number_format = "0.0000"
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
            cell.fill = RES_FILL
    for c, wd in zip(range(1, 7), [26, 16, 14, 16, 14, 18]):
        ws.column_dimensions[get_column_letter(c)].width = wd
    ws.row_dimensions[1].height = 42
    return ws


def excel(N):
    wb = Workbook()
    wb.remove(wb.active)
    hoja_ejemplo(wb, N)
    hoja_runs(wb, N)
    path = os.path.join(OUT, "fila_banco_%d_clientes.xlsx" % N)
    wb.save(path)
    print("  escrito", os.path.basename(path))


# ---------------------------------------------------------------- main
if __name__ == "__main__":
    print("Tablas LaTeX:")
    for N in (20, 200):
        w("tbl%d.tex" % N, tabla_detalle(N))
        w("meas%d.tex" % N, tabla_medidas(N))
        w("runs%d.tex" % N, tabla_corridas(N))
        w("ci%d.tex" % N, tabla_ic(N))
    w("resumen.tex", tabla_resumen())
    w("cv.tex", tabla_cv())
    w("valgen.tex", tabla_validacion())
    print("Figuras:")
    figura_dispersion()
    figura_precision()
    print("Excel:")
    excel(20)
    excel(200)
    print("\nListo.")
